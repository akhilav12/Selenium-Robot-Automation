import openpyxl as py
"""
To Read the Excel sheet rows and column and append it in a dictionary
"""

def ExteranetExcelRead(path):
    """
        This Function will Read Excel from mentioned path.
        Args:
        Path of Excel file(String)
        Returns:
        This function will return List object
    """
    wb=py.load_workbook(path)
    print(type(wb))
    sheet=wb.active
    rows=sheet.max_row
    columns=sheet.max_column
    print(rows)
    print(columns)
    print("For loop started")
    list=[]
    for r in range(2,rows+1):
        #values = []
        dict = {}
        for j in range(1,columns+1):
            cell_obj=sheet.cell(row=r,column=j)
            dict[sheet.cell(row=1,column=j).value] = cell_obj.value
        print("Dictionary is created")
        list.append(dict)
    print(len(list))
    wb.close()
    return list
